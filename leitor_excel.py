from statistics import median

import openpyxl
import pandas as pd

from gerador_excel import COL_META

COL_META_IDX = openpyxl.utils.column_index_from_string(COL_META)

LIMITE_BARATO = 0.5   # abaixo de 50% da mediana -> alerta
LIMITE_CARO = 2.0     # acima de 200% da mediana -> alerta

COLS_CALCULADAS = [
    "Qtd", "Menor Preço (R$)", "Subtotal Otimizado (R$)",
    "Fornecedor Vencedor", "Fornecedor Escolhido", "Alerta",
]


# ─────────────────────────────────────────── leitura de preço (requisito 3)
def limpar_valor_monetario(val):
    """
    Converte qualquer entrada em float:
      ''  None  ' '        -> None (célula vazia)
      'R$ 10,50'           -> 10.50
      '10,50' / '10.50'    -> 10.50
      '1.234,56'           -> 1234.56  (padrão brasileiro)
      '1,234.56'           -> 1234.56  (padrão americano)
      10.5 (número)        -> 10.50
    Valores <= 0 são tratados como não preenchidos (None).
    """
    if val is None or isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        return float(val) if (not pd.isna(val) and val > 0) else None

    s = str(val).upper().replace("R$", "").replace(" ", "").replace("\xa0", "").strip()
    if not s or s in ("NAN", "-", "—"):
        return None

    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):        # 1.234,56
            s = s.replace(".", "").replace(",", ".")
        else:                                   # 1,234.56
            s = s.replace(",", "")
    elif "," in s:                              # 10,50
        s = s.replace(".", "").replace(",", ".")

    try:
        v = float(s)
    except ValueError:
        return None
    return v if v > 0 else None


# ─────────────────────────────────────────── localização da tabela na aba
def _localizar_cabecalho(ws, max_linhas=12):
    for linha in range(1, min(max_linhas, ws.max_row) + 1):
        col_med = col_preco = col_qtd = None
        for col in range(1, min(ws.max_column, COL_META_IDX) + 1):
            txt = str(ws.cell(row=linha, column=col).value or "").strip().lower()
            if not txt:
                continue
            if col_med is None and any(x in txt for x in ("medicamento", "descri", "produto")):
                col_med = col
            elif col_preco is None and any(x in txt for x in ("valor", "preço", "preco", "unitario", "unitário")):
                col_preco = col
            elif col_qtd is None and any(x in txt for x in ("qtd", "quant")):
                col_qtd = col
        if col_med and col_preco:
            return linha, col_med, col_preco, col_qtd
    return None, None, None, None


def _nome_arquivo(arquivo):
    return str(getattr(arquivo, "name", arquivo)).split("/")[-1].split("\\")[-1]


def _forn_pelo_arquivo(arquivo):
    return (_nome_arquivo(arquivo)
            .replace("Cotacao_", "").replace(".xlsx", "").replace(".xls", "")
            .replace("_", " ").strip())


# ─────────────────────────────────────────── comparativo (requisitos 2 e 3)
def processar_comparativo(arquivos_enviados, quantidades_padrao=None):
    """
    Lê as planilhas devolvidas e monta o comparativo.
    Retorna (df, relatorio). O comparativo NÃO decide empates — só os detecta.
    """
    relatorio = {"lidos": [], "erros": [], "avisos": []}
    quantidades_padrao = quantidades_padrao or {}
    if not arquivos_enviados:
        return None, relatorio

    dados, quantidades, vistos = {}, {}, {}

    for arquivo in arquivos_enviados:
        nome_arq = _nome_arquivo(arquivo)
        try:
            if hasattr(arquivo, "seek"):
                arquivo.seek(0)
            ws = openpyxl.load_workbook(arquivo, data_only=True).active
        except Exception as e:
            relatorio["erros"].append(f"{nome_arq}: não foi possível abrir ({e}).")
            continue

        forn_meta = ws[f"{COL_META}1"].value
        if forn_meta and str(forn_meta).strip():
            forn = str(forn_meta).strip()
        else:
            forn = _forn_pelo_arquivo(arquivo)
            relatorio["avisos"].append(
                f"{nome_arq}: sem identificação interna — assumi \"{forn}\" pelo nome do arquivo.")

        if forn in vistos:
            relatorio["avisos"].append(
                f"{nome_arq}: \"{forn}\" já veio em {vistos[forn]}; o mais recente substituiu.")
        vistos[forn] = nome_arq

        linha_cab, col_med, col_preco, col_qtd_vis = _localizar_cabecalho(ws)
        if not linha_cab:
            relatorio["erros"].append(
                f"{nome_arq} ({forn}): não achei as colunas de medicamento e valor. Ficou de fora.")
            continue

        itens_lidos = precos_lidos = 0
        sem_preco = []

        for linha in range(linha_cab + 1, ws.max_row + 1):
            med = str(ws.cell(row=linha, column=col_med).value or "").strip().upper()
            if not med or med == "NAN":
                continue
            itens_lidos += 1
            preco = limpar_valor_monetario(ws.cell(row=linha, column=col_preco).value)
            if preco is None:
                sem_preco.append(med)
            else:
                precos_lidos += 1

            dados.setdefault(med, {})[forn] = preco

            qtd = None
            meta = ws.cell(row=linha, column=COL_META_IDX).value
            if meta is not None:
                try:
                    qtd = int(float(meta))
                except (TypeError, ValueError):
                    qtd = None
            if qtd is None and col_qtd_vis:
                try:
                    qtd = int(float(ws.cell(row=linha, column=col_qtd_vis).value))
                except (TypeError, ValueError):
                    qtd = None
            if qtd is None:
                qtd = quantidades_padrao.get(med)
            if not qtd or qtd < 1:
                qtd = 1
            quantidades[med] = max(quantidades.get(med, 1), int(qtd))

        if itens_lidos == 0:
            relatorio["erros"].append(f"{nome_arq} ({forn}): nenhum item legível. Ficou de fora.")
            continue
        if precos_lidos == 0:
            relatorio["avisos"].append(f"{forn}: {itens_lidos} itens, nenhum preço preenchido.")
        elif sem_preco:
            relatorio["avisos"].append(f"{forn}: {len(sem_preco)} de {itens_lidos} itens sem preço.")

        relatorio["lidos"].append(
            {"fornecedor": forn, "arquivo": nome_arq, "itens": itens_lidos, "precos": precos_lidos})

    if not dados:
        return None, relatorio

    df = pd.DataFrame.from_dict(dados, orient="index").sort_index()
    df.index.name = "Medicamento"
    cols_forn = sorted(df.columns.tolist())
    df = df[cols_forn]

    df["Menor Preço (R$)"] = df[cols_forn].min(axis=1)
    df["Fornecedor Vencedor"] = df.apply(lambda r: detectar_vencedor(r, cols_forn), axis=1)
    df["Qtd"] = [quantidades.get(m, 1) for m in df.index]
    df["Subtotal Otimizado (R$)"] = df["Menor Preço (R$)"] * df["Qtd"]
    df["Alerta"] = df.apply(lambda r: _alertas(r, cols_forn), axis=1)

    ordem = (["Qtd"] + cols_forn +
             ["Menor Preço (R$)", "Subtotal Otimizado (R$)", "Fornecedor Vencedor", "Alerta"])
    return df[ordem], relatorio


def colunas_de_fornecedores(df):
    return [c for c in df.columns if c not in COLS_CALCULADAS]


# ─────────────────────────────────────────── empates (requisito 1)
def detectar_vencedor(row, cols_forn):
    """Retorna o fornecedor de menor preço, ou 'EMPATE (a / b)' quando há mais de um."""
    menor = row.get("Menor Preço (R$)")
    if menor is None or pd.isna(menor):
        return "Nenhum"
    venc = []
    for col in cols_forn:
        v = row.get(col)
        if v is None or pd.isna(v):
            continue
        try:
            if abs(float(v) - float(menor)) < 0.001:
                venc.append(col)
        except (TypeError, ValueError):
            continue
    if len(venc) > 1:
        return f"EMPATE ({' / '.join(venc)})"
    if len(venc) == 1:
        return venc[0]
    return "Nenhum"


def listar_empates(df):
    """
    Retorna [{medicamento, preco, candidatos:[...]}] só dos itens realmente empatados.
    Base para a tela de desempate.
    """
    empates = []
    if df is None or "Fornecedor Vencedor" not in df.columns:
        return empates
    for med, row in df.iterrows():
        venc = str(row["Fornecedor Vencedor"])
        if venc.startswith("EMPATE"):
            candidatos = venc.replace("EMPATE (", "").replace(")", "").split(" / ")
            empates.append({
                "medicamento": med,
                "preco": float(row["Menor Preço (R$)"]),
                "candidatos": candidatos,
            })
    return empates


def aplicar_escolhas(df, escolhas):
    """
    Recebe {medicamento: fornecedor_escolhido} e devolve um novo df com a coluna
    'Fornecedor Escolhido' resolvida. Não apaga os preços dos concorrentes.
    """
    novo = df.copy()
    novo["Fornecedor Escolhido"] = novo["Fornecedor Vencedor"]
    for med, row in novo.iterrows():
        venc = str(row["Fornecedor Vencedor"])
        if venc.startswith("EMPATE"):
            escolhido = escolhas.get(med)
            novo.loc[med, "Fornecedor Escolhido"] = escolhido if escolhido else venc
    return novo


# ─────────────────────────────────────────── alertas de preço
def _alertas(row, cols_forn):
    precos = {}
    for col in cols_forn:
        v = row.get(col)
        if v is None or pd.isna(v):
            continue
        try:
            precos[col] = float(v)
        except (TypeError, ValueError):
            continue
    if not precos:
        return "Ninguém cotou"
    if len(precos) == 1:
        return f"Só {list(precos)[0]} cotou"

    partes = []
    if len(precos) >= 3:
        med = median(precos.values())
        if med > 0:
            baratos = [c for c, v in precos.items() if v < med * LIMITE_BARATO]
            caros = [c for c, v in precos.items() if v > med * LIMITE_CARO]
            if baratos:
                partes.append(f"Muito barato: {', '.join(baratos)} — confira")
            if caros:
                partes.append(f"Muito caro: {', '.join(caros)}")
    faltando = [c for c in cols_forn if c not in precos]
    if faltando:
        partes.append(f"Sem preço: {', '.join(faltando)}")
    return " | ".join(partes)
