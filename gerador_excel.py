import io
import zipfile

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side

# Coluna técnica, oculta e protegida, usada só pelo sistema.
# Z1 = nome do fornecedor | Z2 = id da cotação | Z4.. = quantidade de cada item.
COL_META = "Z"
SENHA_PROTECAO = "cotacao"

FONTE = "Arial"


def gerar_zip_cotacoes(itens, fornecedores_selecionados, cotacao_id=None):
    """
    Gera um .xlsx por fornecedor, todos dentro de um .zip.

    itens: lista de dicts {"medicamento": str, "qtd": int}.
           Também aceita uma lista simples de strings (quantidade vira 1).
    O fornecedor vê apenas Item / Descrição / Preço Unitário — a quantidade
    viaja na coluna oculta e volta junto com o arquivo preenchido.
    """
    itens_norm = _normalizar_itens(itens)
    buffer_zip = io.BytesIO()

    with zipfile.ZipFile(buffer_zip, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for forn in fornecedores_selecionados:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Cotação"

            fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            font_header = Font(name=FONTE, size=11, bold=True, color="FFFFFF")
            fill_input = PatternFill(start_color="FFF9DB", end_color="FFF9DB", fill_type="solid")
            border_thin = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9'),
            )

            ws['A1'] = f"COTAÇÃO DE PREÇOS - FORNECEDOR: {str(forn).upper()}"
            ws['A1'].font = Font(name=FONTE, size=14, bold=True, color="1F4E78")

            ws['A2'] = ("Preencha somente a coluna C (Preço Unitário), destacada em amarelo. "
                        "Não altere, apague ou reordene as linhas — devolva o arquivo no mesmo formato.")
            ws['A2'].font = Font(name=FONTE, size=9, italic=True, color="808080")

            headers = ["Item", "Descrição do Medicamento", "Preço Unitário (R$)"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num, value=header)
                cell.fill = fill_header
                cell.font = font_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.protection = Protection(locked=True)

            for idx, item in enumerate(itens_norm, start=1):
                row_idx = idx + 3

                c1 = ws.cell(row=row_idx, column=1, value=idx)
                c2 = ws.cell(row=row_idx, column=2, value=item["medicamento"])
                c3 = ws.cell(row=row_idx, column=3, value=None)

                c1.alignment = Alignment(horizontal="center")
                c2.alignment = Alignment(horizontal="left")
                c3.alignment = Alignment(horizontal="right")
                c3.number_format = 'R$ #,##0.00'
                c3.fill = fill_input

                for c in (c1, c2, c3):
                    c.font = Font(name=FONTE, size=10)
                    c.border = border_thin

                c1.protection = Protection(locked=True)
                c2.protection = Protection(locked=True)
                c3.protection = Protection(locked=False)

                # Quantidade oculta — o fornecedor não vê
                ws[f"{COL_META}{row_idx}"] = int(item["qtd"])

            # Metadados técnicos
            ws[f"{COL_META}1"] = str(forn)
            ws[f"{COL_META}2"] = "" if cotacao_id is None else str(cotacao_id)

            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions[COL_META].hidden = True

            # Protege a planilha: sem senha, qualquer um reexibe a coluna oculta
            ws.protection.set_password(SENHA_PROTECAO)
            ws.protection.sheet = True
            ws.protection.formatColumns = True

            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            nome_arquivo = f"Cotacao_{str(forn).replace(' ', '_')}.xlsx"
            zip_file.writestr(nome_arquivo, excel_buffer.getvalue())

    buffer_zip.seek(0)
    return buffer_zip


def _normalizar_itens(itens):
    """Aceita lista de strings ou lista de dicts e devolve sempre dicts."""
    normalizados = []
    for item in itens:
        if isinstance(item, dict):
            med = str(item.get("medicamento", "")).strip()
            try:
                qtd = max(1, int(item.get("qtd", 1)))
            except (TypeError, ValueError):
                qtd = 1
        else:
            med = str(item).strip()
            qtd = 1
        if med:
            normalizados.append({"medicamento": med, "qtd": qtd})
    return normalizados
