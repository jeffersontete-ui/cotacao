from statistics import median

import openpyxl
import pandas as pd

from gerador_excel import COL_META

COL_META_IDX = openpyxl.utils.column_index_from_string(COL_META)

# Um preço abaixo deste múltiplo da mediana (ou acima do outro) vira alerta.
LIMITE_BARATO = 0.5
LIMITE_CARO = 2.0


def limpar_valor_monetario(val):
    """Converte para float valores como 'R$ 3,00', '3.00', '1.000,00' ou 1000.00."""
    if val is None:
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        return float(val) if not pd.isna(val) else None

    val_str = str(val).upper().replace("R$", "").replace(" ", "").strip()
    if not val_str or val_str == "NAN":
        return None

    if "," in val_str and "." in val_str:
        # 1.000,00 (brasileiro) vs 1,000.00 (americano)
        if val_str.rfind(",") > val_str.rfind("."):
            val_str = val_str.replace(".", "").replace(",", ".")
        else:
            val_str = val_str.replace(",", "")
    elif "," in val_str:
        val_str = val_str.replace(",", ".")

    try:
        valor = float(val_str)
    except ValueError:
        return None

    return valor if valor > 0 else None


def _localizar_cabecalho(ws, max_linhas=12):
    """
    Procura nas primeiras linhas a que contém as colunas de medicamento e preço.
    Retorna (linha_cabecalho, col_medicamento, col_preco, col_qtd_visivel).
    """
    for linha in range(1, min(max_linhas, ws.max_row) + 1):
        col_med = col_preco = col_qtd = None
        for col in range(1, min(ws.max_column, COL_META_IDX) + 1):
            texto = str(ws.cell(row=linha, column=col).value or "").strip().lower()
            if not texto:
                continue
            if col_med is None and any(x in texto for x in ("medicamento", "descri", "produto")):
                col_med = col
            elif col_preco is None and any(x in texto for x in ("preço", "preco", "valor", "unitario", "unitário")):
                col_preco = col
            elif col_qtd is None and any(x in texto for x in ("qtd", "quant")):
                col_qtd = col
        if col_med and col_preco:
            return linha, col_med, col_preco, col_qtd
    return None, None, None, None


def _nome_do_arquivo(arquivo):
    nome = getattr(arquivo, "name", str(arquivo))
    return str(nome).split("/")[-1].split("\\")[-1]


def _fornecedor_pelo_arquivo(arquivo):
    return (
        _nome_do_arquivo(arquivo)
        .replace("Cotacao_", "")
        .replace(".xlsx", "")
        .replace(".xls", "")
        .replace("_", " ")
        .strip()
    )


def processar_comparativo(arquivos_enviados, quantidades_padrao=None):
    """
    Lê as planilhas devolvidas pelos fornecedores e monta o mapa comparativo.

    Retorna (df, relatorio). O relatorio traz:
      lidos  -> [{"fornecedor", "arquivo", "itens", "precos"}]
      erros  -> mensagens de arquivos que não puderam ser usados
      avisos -> problemas que não impedem a comparação
    """
    relatorio = {"lidos": [], "erros": [], "avisos": []}
    quantidades_padrao = quantidades_padrao or {}

    if not arquivos_enviados:
        return None, relatorio

    dados_medicamentos = {}
    quantidades = {}
    fornecedores_vistos = {}

    for arquivo in arquivos_enviados:
        nome_arquivo = _nome_do_arquivo(arquivo)
        try:
            if hasattr(arquivo, "seek"):
                arquivo.seek(0)
            wb = openpyxl.load_workbook(arquivo, data_only=True, read_only=False)
            ws = wb.active
        except Exception as e:
            relatorio["erros"].append(f"{nome_arquivo}: não foi possível abrir o arquivo ({e}).")
            continue

        # 1) Quem é o fornecedor: primeiro a célula técnica, depois o nome do arquivo
        forn_meta = ws[f"{COL_META}1"].value
        if forn_meta and str(forn_meta).strip():
            nome_fornecedor = str(forn_meta).strip()
        else:
            nome_fornecedor = _fornecedor_pelo_arquivo(arquivo)
            relatorio["avisos"].append(
                f"{nome_arquivo}: identificação interna ausente — fornecedor deduzido do nome do "
                f"arquivo como \"{nome_fornecedor}\". Confira se está correto."
            )

        if nome_fornecedor in fornecedores_vistos:
            relatorio["avisos"].append(
                f"{nome_arquivo}: \"{nome_fornecedor}\" já havia sido carregado em "
                f"{fornecedores_vistos[nome_fornecedor]}. O arquivo mais recente substituiu o anterior."
            )
        fornecedores_vistos[nome_fornecedor] = nome_arquivo

        # 2) Onde está a tabela
        linha_cab, col_med, col_preco, col_qtd_visivel = _localizar_cabecalho(ws)
        if not linha_cab:
            relatorio["erros"].append(
                f"{nome_arquivo} ({nome_fornecedor}): não encontrei as colunas de medicamento e "
                f"preço. O arquivo NÃO entrou na comparação."
            )
            continue

        itens_lidos = 0
        precos_lidos = 0
        sem_preco = []

        for linha in range(linha_cab + 1, ws.max_row + 1):
            med = str(ws.cell(row=linha, column=col_med).value or "").strip().upper()
            if not med or med == "NAN":
                continue

            itens_lidos += 1
            preco = limpar_valor_monetario(ws.cell(row=linha, column=col_preco).value)
            if preco is None:
                sem_preco.append(med)
            else:
                precos_lidos += 1

            dados_medicamentos.setdefault(med, {})
            dados_medicamentos[med][nome_fornecedor] = preco

            # 3) Quantidade: coluna oculta > coluna visível > cotação salva > 1
            qtd = None
            valor_meta = ws.cell(row=linha, column=COL_META_IDX).value
            if valor_meta is not None:
                try:
                    qtd = int(float(valor_meta))
                except (TypeError, ValueError):
                    qtd = None
            if qtd is None and col_qtd_visivel:
                try:
                    qtd = int(float(ws.cell(row=linha, column=col_qtd_visivel).value))
                except (TypeError, ValueError):
                    qtd = None
            if qtd is None:
                qtd = quantidades_padrao.get(med)
            if qtd is None or qtd < 1:
                qtd = 1

            quantidades[med] = max(quantidades.get(med, 1), int(qtd))

        if itens_lidos == 0:
            relatorio["erros"].append(
                f"{nome_arquivo} ({nome_fornecedor}): nenhum item legível. NÃO entrou na comparação."
            )
            continue

        if precos_lidos == 0:
            relatorio["avisos"].append(
                f"{nome_fornecedor}: {itens_lidos} itens lidos, mas nenhum preço preenchido."
            )
        elif sem_preco:
            relatorio["avisos"].append(
                f"{nome_fornecedor}: {len(sem_preco)} de {itens_lidos} itens vieram sem preço."
            )

        relatorio["lidos"].append({
            "fornecedor": nome_fornecedor,
            "arquivo": nome_arquivo,
            "itens": itens_lidos,
            "precos": precos_lidos,
        })

    if not dados_medicamentos:
        return None, relatorio

    df = pd.DataFrame.from_dict(dados_medicamentos, orient="index")
    df.index.name = "Medicamento"
    df = df.sort_index()

    cols_fornecedores = sorted(df.columns.tolist())
    df = df[cols_fornecedores]

    df["Menor Preço (R$)"] = df[cols_fornecedores].min(axis=1)
    df["Fornecedor Vencedor"] = df.apply(
        lambda row: identificar_vencedor(row, cols_fornecedores), axis=1
    )
    df["Qtd"] = [quantidades.get(m, 1) for m in df.index]
    df["Subtotal Otimizado (R$)"] = df["Menor Preço (R$)"] * df["Qtd"]
    df["Alerta"] = df.apply(lambda row: _alertas_da_linha(row, cols_fornecedores), axis=1)

    cols_finais = (
        ["Qtd"] + cols_fornecedores
        + ["Menor Preço (R$)", "Subtotal Otimizado (R$)", "Fornecedor Vencedor", "Alerta"]
    )
    return df[cols_finais], relatorio


def identificar_vencedor(row, cols_fornecedores):
    """Retorna o fornecedor de menor preço, ou EMPATE quando há mais de um."""
    menor_valor = row.get("Menor Preço (R$)")
    if menor_valor is None or pd.isna(menor_valor):
        return "Nenhum"

    vencedores = []
    for col in cols_fornecedores:
        val = row.get(col)
        if val is None or pd.isna(val):
            continue
        try:
            if abs(float(val) - float(menor_valor)) < 0.001:
                vencedores.append(col)
        except (TypeError, ValueError):
            continue

    if len(vencedores) > 1:
        return f"EMPATE ({' / '.join(vencedores)})"
    if len(vencedores) == 1:
        return vencedores[0]
    return "Nenhum"


def _alertas_da_linha(row, cols_fornecedores):
    """Sinaliza itens que merecem conferência antes de fechar o pedido."""
    precos = {}
    for col in cols_fornecedores:
        val = row.get(col)
        if val is None or pd.isna(val):
            continue
        try:
            precos[col] = float(val)
        except (TypeError, ValueError):
            continue

    if not precos:
        return "Ninguém cotou"
    if len(precos) == 1:
        return f"Só {list(precos)[0]} cotou"

    alertas = []
    if len(precos) >= 3:
        mediana = median(precos.values())
        if mediana > 0:
            baratos = [c for c, v in precos.items() if v < mediana * LIMITE_BARATO]
            caros = [c for c, v in precos.items() if v > mediana * LIMITE_CARO]
            if baratos:
                alertas.append(f"Preço muito abaixo da mediana: {', '.join(baratos)} — confira")
            if caros:
                alertas.append(f"Preço muito acima da mediana: {', '.join(caros)}")

    faltando = [c for c in cols_fornecedores if c not in precos]
    if faltando:
        alertas.append(f"Sem preço: {', '.join(faltando)}")

    return " | ".join(alertas)
