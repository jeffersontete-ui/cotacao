import io
import zipfile

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

FONTE = "Arial"
COL_ESCOLHIDO = "Fornecedor Escolhido"


def preparar_pedidos(df_comparativo):
    """
    Normaliza o comparativo para a geração de pedidos.

    Usa a coluna "Fornecedor Escolhido" quando existir (desempate feito na tela);
    caso contrário usa "Fornecedor Vencedor".
    Retorna (df_fechados, df_pendentes).
    """
    df = df_comparativo.reset_index() if df_comparativo.index.name == "Medicamento" else df_comparativo.copy()

    coluna = COL_ESCOLHIDO if COL_ESCOLHIDO in df.columns else "Fornecedor Vencedor"
    df["_forn"] = df[coluna].astype(str)

    pendente = (
        df["_forn"].str.startswith("EMPATE", na=False)
        | df["_forn"].isin(["Nenhum", "nan", "None", ""])
        | df["Menor Preço (R$)"].isna()
    )

    if "Qtd" not in df.columns:
        df["Qtd"] = 1
    if "Subtotal Otimizado (R$)" not in df.columns:
        df["Subtotal Otimizado (R$)"] = df["Menor Preço (R$)"] * df["Qtd"]

    return df[~pendente].copy(), df[pendente].copy()


def gerar_zip_pedidos(df_comparativo):
    """Gera um .xlsx de pedido de compra por fornecedor vencedor, dentro de um .zip."""
    buffer_zip = io.BytesIO()
    df_fechados, _ = preparar_pedidos(df_comparativo)

    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font_header = Font(name=FONTE, size=11, bold=True, color="FFFFFF")
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    with zipfile.ZipFile(buffer_zip, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for forn, df_itens in df_fechados.groupby("_forn"):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Pedido de Compra"

            ws['A1'] = f"PEDIDO DE COMPRA - FORNECEDOR: {str(forn).upper()}"
            ws['A1'].font = Font(name=FONTE, size=14, bold=True, color="1F4E78")

            headers = ["Item", "Descrição do Medicamento", "Qtd", "Preço Unitário (R$)", "Subtotal (R$)"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num, value=header)
                cell.fill = fill_header
                cell.font = font_header
                cell.alignment = Alignment(horizontal="center", vertical="center")

            primeira_linha = 4
            for idx, (_, row) in enumerate(df_itens.iterrows(), start=1):
                r = idx + 3
                qtd = int(row.get("Qtd", 1) or 1)
                preco = float(row["Menor Preço (R$)"])

                ws.cell(row=r, column=1, value=idx).alignment = Alignment(horizontal="center")
                ws.cell(row=r, column=2, value=str(row["Medicamento"]))
                ws.cell(row=r, column=3, value=qtd).alignment = Alignment(horizontal="center")

                c_preco = ws.cell(row=r, column=4, value=preco)
                c_preco.number_format = 'R$ #,##0.00'

                # Subtotal como fórmula: a planilha recalcula se você ajustar a quantidade
                c_sub = ws.cell(row=r, column=5, value=f"=C{r}*D{r}")
                c_sub.number_format = 'R$ #,##0.00'

                for col in range(1, 6):
                    c = ws.cell(row=r, column=col)
                    c.border = border_thin
                    if c.font.name != FONTE:
                        c.font = Font(name=FONTE, size=10)

            ultima_linha = len(df_itens) + 3
            linha_total = ultima_linha + 1

            lbl = ws.cell(row=linha_total, column=4, value="TOTAL DO PEDIDO:")
            lbl.font = Font(name=FONTE, size=11, bold=True)
            lbl.alignment = Alignment(horizontal="right")

            tot = ws.cell(row=linha_total, column=5,
                          value=f"=SUM(E{primeira_linha}:E{ultima_linha})")
            tot.font = Font(name=FONTE, size=11, bold=True, color="1F4E78")
            tot.number_format = 'R$ #,##0.00'

            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20

            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            zip_file.writestr(
                f"Pedido_Compra_{str(forn).replace(' ', '_')}.xlsx",
                excel_buffer.getvalue(),
            )

    buffer_zip.seek(0)
    return buffer_zip
