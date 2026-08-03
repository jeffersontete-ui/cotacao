import io
import json
import os
import zipfile

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side

# Configuração da página (Modo amplo para facilitar visualização)
st.set_page_config(
    page_title="Gestão de Cotações", page_icon="💊", layout="wide"
)

# -----------------------------------------------------------------------------
# FUNÇÕES DE PERSISTÊNCIA EM JSON
# -----------------------------------------------------------------------------
ARQUIVO_JSON = "fornecedores.json"


def carregar_fornecedores():
  """Lê os fornecedores salvos no arquivo JSON."""
  if os.path.exists(ARQUIVO_JSON):
    try:
      with open(ARQUIVO_JSON, "r", encoding="utf-8") as f:
        return json.load(f)
    except Exception:
      return []
  return []


def salvar_fornecedores(dados):
  """Salva a lista de fornecedores no arquivo JSON."""
  with open(ARQUIVO_JSON, "w", encoding="utf-8") as f:
    json.dump(dados, f, indent=2, ensure_ascii=False)


def gerar_zip_cotacoes(medicamentos, fornecedores_selecionados):
  buffer_zip = io.BytesIO()

  with zipfile.ZipFile(buffer_zip, "w", zipfile.ZIP_DEFLATED) as zip_file:
    for forn in fornecedores_selecionados:
      wb = openpyxl.Workbook()
      ws = wb.active
      ws.title = "Cotação"

      # ATIVAR PROTEÇÃO DA PLANILHA
      ws.protection.sheet = True

      # Estilos do cabeçalho
      fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
      font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
      border_thin = Border(
          left=Side(style='thin', color='D9D9D9'),
          right=Side(style='thin', color='D9D9D9'),
          top=Side(style='thin', color='D9D9D9'),
          bottom=Side(style='thin', color='D9D9D9')
      )

      # Título do Documento
      ws['A1'] = f"COTAÇÃO DE PREÇOS - FORNECEDOR: {forn.upper()}"
      ws['A1'].font = Font(name="Calibri", size=14, bold=True, color="1F4E78")

      # Cabeçalho da Tabela - APENAS 3 COLUNAS
      headers = ["Item", "Descrição do Medicamento", "Preço Unitário (R$)"]
      ws.append([])
      ws.append(headers)

      for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.protection = Protection(locked=True)

      for idx, med in enumerate(medicamentos, start=1):
        row_idx = idx + 3

        c1 = ws.cell(row=row_idx, column=1, value=idx)
        c2 = ws.cell(row=row_idx, column=2, value=med.strip())
        c3 = ws.cell(row=row_idx, column=3, value=None)

        c1.alignment = Alignment(horizontal="center")
        c2.alignment = Alignment(horizontal="left")
        c3.alignment = Alignment(horizontal="right")
        c3.number_format = 'R$ #,##0.00'

        c1.protection = Protection(locked=True)
        c2.protection = Protection(locked=True)
        c3.protection = Protection(locked=False)

        for c in (c1, c2, c3):
          c.border = border_thin

      ws.column_dimensions['A'].width = 10
      ws.column_dimensions['B'].width = 50
      ws.column_dimensions['C'].width = 25

      excel_buffer = io.BytesIO()
      wb.save(excel_buffer)
      excel_buffer.seek(0)

      nome_arquivo = f"Cotacao_{forn.replace(' ', '_')}.xlsx"
      zip_file.writestr(nome_arquivo, excel_buffer.getvalue())

  buffer_zip.seek(0)
  return buffer_zip


def processar_comparativo(arquivos_carregados):
  dados = []

  for arq in arquivos_carregados:
    try:
      wb = openpyxl.load_workbook(arq, data_only=True)
      ws = wb.active

      titulo_a1 = str(ws['A1'].value or '')
      if "FORNECEDOR:" in titulo_a1:
        nome_fornecedor = titulo_a1.split("FORNECEDOR:")[-1].strip()
      else:
        nome_fornecedor = arq.name.replace(".xlsx", "").replace("Cotacao_", "").replace("_", " ")

      arq.seek(0)
      df = pd.read_excel(arq, skiprows=2)
      df.columns = [str(c).strip() for c in df.columns]

      col_med = [c for c in df.columns if "Descrição" in c or "Medicamento" in c][0]
      col_prc = [c for c in df.columns if "Preço" in c or "Unitário" in c][0]

      df_temp = df[[col_med, col_prc]].copy()
      df_temp.columns = ['Medicamento', 'Preço Unitário']
      df_temp['Fornecedor'] = nome_fornecedor
      df_temp['Preço Unitário'] = pd.to_numeric(df_temp['Preço Unitário'], errors='coerce')
      df_temp = df_temp.dropna(subset=['Medicamento'])

      dados.append(df_temp)
    except Exception as e:
      st.error(f"Erro ao processar o arquivo {arq.name}: {e}")

  if not dados:
    return None

  df_todos = pd.concat(dados, ignore_index=True)
  matriz = df_todos.pivot_table(
      index='Medicamento',
      columns='Fornecedor',
      values='Preço Unitário',
      aggfunc='first'
  )
  matriz['Menor Preço (R$)'] = matriz.min(axis=1)
  matriz['Fornecedor Vencedor'] = matriz.drop(columns=['Menor Preço (R$)']).idxmin(axis=1)

  return matriz


# -----------------------------------------------------------------------------
# NAVEGAÇÃO POR ABAS
# -----------------------------------------------------------------------------
st.title("💊 Sistema Integrado de Cotações e Fornecedores")

aba_cotacao, aba_comparativo, aba_cadastro = st.tabs(
    ["📝 1. Criar Cotação", "📊 2. Comparar Preços", "🏢 3. Cadastro de Fornecedores"]
)

# -----------------------------------------------------------------------------
# ABA 1: CRIAR COTAÇÃO
# -----------------------------------------------------------------------------
with aba_cotacao:
  st.header("Gerar Nova Cotação")

  # 1. Criamos a divisão em duas colunas (60% e 40%)
  col1, col2 = st.columns([3, 2])

  # 2. Tudo dentro de 'with col1:' vai para a esquerda
  with col1:
    lista_padrao = (
        "ACEBROFILINA 25 MG XPE PED 120 ML\nACECLOFENACO 100 MG C/ 12 CPR"
    )
    lista_meds = st.text_area(
        "Lista de Medicamentos (um por linha):",
        value=lista_padrao,
        height=220,
        help="Digite ou cole um medicamento por linha.",
    )

  # 3. Tudo dentro de 'with col2:' vai para a direita
  with col2:
    fornecedores = carregar_fornecedores()
    todos_nomes = [f["distribuidora"] for f in fornecedores if "distribuidora" in f]
    ativos_nomes = [
        f["distribuidora"]
        for f in fornecedores
        if f.get("ativo") and "distribuidora" in f
    ]

    fornecedores_sel = st.multiselect(
        "Selecione os fornecedores para esta cotação:",
        options=todos_nomes,
        default=ativos_nomes,
        help="Fornecedores marcados como Ativos no cadastro vêm pré-selecionados automaticamente.",
    )

  # 3. Botão para Gerar Planilhas
  if st.button("⚙️ Gerar Planilhas de Cotação", use_container_width=True):
    lista_meds_limpa = [m for m in lista_meds.split("\n") if m.strip()]

    if not lista_meds_limpa:
      st.warning("⚠️ Por favor, insira pelo menos um medicamento.")
    elif not fornecedores_sel:
      st.warning("⚠️ Por favor, selecione pelo menos um fornecedor.")
    else:
      zip_data = gerar_zip_cotacoes(lista_meds_limpa, fornecedores_sel)

      st.success("✅ Planilhas geradas com sucesso!")

      st.download_button(
          label="📥 Baixar Arquivos de Cotação (.zip)",
          data=zip_data,
          file_name="Cotacoes_Fornecedores.zip",
          mime="application/zip",
          use_container_width=True,
      )

# -----------------------------------------------------------------------------
# ABA 2: COMPARAR PREÇOS
# -----------------------------------------------------------------------------
with aba_comparativo:
  st.header("📊 Comparativo de Preços das Cotações")
  st.write("Envie abaixo as planilhas Excel preenchidas que você recebeu dos fornecedores:")

  arquivos_enviados = st.file_uploader(
      "Selecione as planilhas dos fornecedores (.xlsx)",
      type=["xlsx"],
      accept_multiple_files=True,
  )

  if arquivos_enviados:
    df_comparativo = processar_comparativo(arquivos_enviados)

    if df_comparativo is not None:
      st.success(f"✅ {len(arquivos_enviados)} planilha(s) analisada(s) com sucesso!")

      st.subheader("📌 Tabela Comparativa de Preços")
      st.dataframe(
          df_comparativo.style.format(
              subset=[c for c in df_comparativo.columns if c != 'Fornecedor Vencedor'],
              formatter="R$ {:.2f}"
          ),
          use_container_width=True,
      )

      total_otimizado = df_comparativo['Menor Preço (R$)'].sum()
      st.metric(
          label="💰 Custo Total Otimizado (Menor Preço de Cada Item)",
          value=f"R$ {total_otimizado:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
      )
    else:
      st.warning("Nenhum dado válido foi encontrado nos arquivos carregados.")

# -----------------------------------------------------------------------------
# ABA 3: CADASTRO DE FORNECEDORES
# -----------------------------------------------------------------------------
with aba_cadastro:
  st.header("Cadastro e Status dos Fornecedores")
  st.caption(
      "Edite diretamente na tabela abaixo. Marque/desmarque a opção 'ativo' e"
      " clique em Salvar."
  )

  # Carrega dados atuais em um DataFrame
  dados_fornecedores = carregar_fornecedores()

  if not dados_fornecedores:
    df_inicial = pd.DataFrame([{
        "distribuidora": "DISTRIBUIDORA EXEMPLO",
        "representante": "Nome do Rep",
        "telefone": "(00) 00000-0000",
        "ativo": True,
    }])
  else:
    df_inicial = pd.DataFrame(dados_fornecedores)

  # Tabela interativa leve e rápida
  df_editado = st.data_editor(
      df_inicial,
      num_rows="dynamic",  # Permite adicionar ➕ ou excluir 🗑️ linhas
      use_container_width=True,
      column_config={
          "distribuidora": st.column_config.TextColumn(
              "Distribuidora", required=True
          ),
          "representante": st.column_config.TextColumn("Representante"),
          "telefone": st.column_config.TextColumn("Telefone"),
          "ativo": st.column_config.CheckboxColumn("Ativo?", default=True),
      },
  )

  # Botão de Salvamento
  if st.button("💾 Salvar Alterações de Fornecedores", type="primary"):
    novos_dados = df_editado.to_dict(orient="records")
    salvar_fornecedores(novos_dados)
    st.success("✅ Cadastro atualizado e salvo com sucesso no arquivo JSON!")